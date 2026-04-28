"""Basic interface for tabular documentation of datasets."""

# pylint: disable=too-many-locals

import csv
import re
from pathlib import Path
from typing import TYPE_CHECKING

from tripper import Triplestore
from tripper.datadoc.context import get_context
from tripper.datadoc.dataset import store, told
from tripper.datadoc.keywords import get_keywords
from tripper.datadoc.utils import addnested, stripnested
from tripper.literal import Literal
from tripper.utils import AttrDict, openfile

if TYPE_CHECKING:  # pragma: no cover
    from typing import Iterable, List, Optional, Protocol, Sequence, Union

    from tripper.datadoc.context import ContextType
    from tripper.datadoc.keywords import KeywordsType

    class Writer(Protocol):
        """Prototype for a class with a write() method."""

        # pylint: disable=too-few-public-methods,missing-function-docstring

        def write(self, data: str) -> None: ...


class TableDoc:
    """Representation of tabular documentation of datasets.

    Arguments:
        headers: Sequence of column header labels.  Nested data can
            be represented by dot-separated label strings (e.g.
            "distribution.downloadURL")
        data: Sequence of rows of data. Each row documents an entry.
        type: Type of data to save (applies to all rows).  Should
            either be one of the pre-defined names: "dataset",
            "distribution", "accessService", "parser" and "generator"
            or an IRI to a class in an ontology.
        theme: Name of one of more themes to load keywords for.
        keywords: Keywords object with additional keywords definitions.
            If not provided, only default keywords are considered.
        context: Additional user-defined context that should be
            returned on top of the default context.  It may be a
            string with an URL to the user-defined context, a dict
            with the user-defined context or a sequence of strings and
            dicts.
        prefixes: Dict with prefixes in addition to those included in the
            JSON-LD context.  Should map namespace prefixes to IRIs.
        strip: Whether to strip leading and trailing whitespaces from cells.
        strict: Whether to raise an `InvalidKeywordError` exception if `d`
            contains an unknown key.
        redefine: Determine how to handle redefinition of existing
            keywords.  Should be one of the following strings:
              - "allow": Allow redefining a keyword. Emits a
                `RedefineKeywordWarning`.
              - "skip": Don't redefine existing keyword. Emits a
                `RedefineKeywordWarning`.
              - "raise": Raise an RedefineError (default).
        baseiri: If given, it will be used as a base iri to
            resolve relative IRIs. (I.e. Not valid URLs).
    """

    # pylint: disable=redefined-builtin,too-few-public-methods
    # pylint: disable=too-many-arguments,too-many-instance-attributes

    def __init__(
        self,
        headers: "Sequence[str]",
        data: "Sequence[Sequence[str]]",
        type: "Optional[str]" = None,
        theme: "Optional[Union[str, Sequence[str]]]" = None,
        keywords: "Optional[KeywordsType]" = None,
        context: "Optional[ContextType]" = None,
        prefixes: "Optional[dict]" = None,
        strip: bool = True,
        strict: bool = False,
        redefine: str = "raise",
        baseiri: "Optional[str]" = None,
    ) -> None:
        if theme is None and keywords is None and context is None:
            theme = "ddoc:datadoc"

        self.headers = list(headers)
        self.data = [list(row) for row in data]
        self.type = type
        self.theme = theme
        self.keywords = get_keywords(
            keywords=keywords,
            theme=theme,
            strict=strict,
            redefine=redefine,
        )
        self.context = get_context(
            context=context,
            keywords=self.keywords,
            default_theme=None,
            prefixes=prefixes,
        )
        self.strip = strip
        self.baseiri = baseiri

    def save(self, ts: Triplestore) -> dict:
        """Save tabular datadocumentation to triplestore.

        Returns a dict with the JSON-LD written to the triplestore.
        """
        self.context.add_context(
            {prefix: str(ns) for prefix, ns in ts.namespaces.items()}
        )

        for prefix, ns in self.context.get_prefixes().items():
            ts.bind(prefix, ns)

        return store(
            ts,
            self.asdicts(),
            type=self.type,
            keywords=self.keywords,
            context=self.context,
            baseiri=self.baseiri,
        )

    def asdicts(self) -> "List[dict]":
        """Return the table as a list of dicts."""
        # Parse column headers. Follow the syntax defined in
        # https://emmc-asbl.github.io/tripper/latest/datadoc/documenting-a-resource/#documenting-as-table
        columns = [
            Column(head, context=self.context, strip=self.strip)
            for head in self.headers
        ]
        results = []
        for row in self.data:
            d = AttrDict()
            for i, col in enumerate(columns):
                cell = row[i].strip() if row[i] and self.strip else row[i]
                col.add(d, cell)
            results.append(stripnested(d))
        ld = told(
            results,
            type=self.type,
            theme=self.theme,
            prefixes=self.context.get_prefixes(),
            context=self.context,
        )
        dicts = ld["@graph"]
        return dicts

    @staticmethod
    def fromdicts(
        dicts: "Sequence[dict]",
        type: "Optional[str]" = None,
        keywords: "Optional[KeywordsType]" = None,
        context: "Optional[ContextType]" = None,
        prefixes: "Optional[dict]" = None,
        strip: bool = True,
    ) -> "TableDoc":
        """Create new TableDoc instance from a sequence of dicts.

        Arguments:
            dicts: Sequence of single-resource dicts.
            type: Type of data to save (applies to all rows).  Should
                either be one of the pre-defined names: "Dataset",
                "Distribution", "AccessService", "Parser" and
                "Generator" or an IRI to a class in an ontology.
            keywords: Keywords object with additional keywords definitions.
                If not provided, only default keywords are considered.
            context: Additional user-defined context that should be
                returned on top of the default context.  It may be a
                string with an URL to the user-defined context, a dict
                with the user-defined context or a sequence of strings
                and dicts.
            prefixes: Dict with prefixes in addition to those included
                in the JSON-LD context.  Should map namespace prefixes
                to IRIs.
            strip: Whether to strip leading and trailing whitespaces
                from cells.

        Returns:
            New TableDoc instance.

        """
        # Store the headers as keys in a dict to keep ordering
        headdict = {"@id": True}

        def addheaddict(d, prefix=""):
            """Add keys in `d` to headdict.

            Nested dicts will result in dot-separated keys.
            """
            for k, v in d.items():
                if k == "@context":
                    pass
                elif isinstance(v, dict):
                    d = v.copy()
                    d.pop("@type", None)
                    addheaddict(d, k + ".")
                else:
                    headdict[prefix + k] = True

        def tolist(v, mult, pad=None):
            """Return `v` as a list of length `mult` with given padding."""
            if isinstance(v, list):
                return v + [pad] * (mult - len(v))
            return [v] + [pad] * (mult - 1)

        # Assign the headdict
        for d in dicts:
            addheaddict(d)

        headers = list(headdict)

        # Calculate multiplicity of each header label
        mult = [1] * len(headers)
        for dct in dicts:
            for i, head in enumerate(headers):
                if head in dct and isinstance(dct[head], list):
                    mult[i] = max(mult[i], len(dct[head]))

        # Assign table data. Multiplicity and nested dicts are accounted for
        data = []
        for dct in dicts:
            row = []
            for head, m in zip(headers, mult):
                if head in dct:
                    row.extend(tolist(dct[head], m))
                else:
                    d = dct
                    for key in head.split("."):
                        d = d.get(key, {})
                    row.extend(tolist(d if d != {} else None, m))
            data.append(row)

        # New multiplied header
        newheaders = []
        for head, m in zip(headers, mult):
            newheaders.extend(tolist(head, m, head))

        return TableDoc(
            headers=newheaders,
            data=data,
            type=type,
            keywords=keywords,
            context=context,
            prefixes=prefixes,
            strip=strip,
        )

    @staticmethod
    def parse_csv(
        csvfile: "Union[Iterable[str], Path, str]",
        type: "Optional[str]" = None,
        keywords: "Optional[KeywordsType]" = None,
        context: "Optional[ContextType]" = None,
        prefixes: "Optional[dict]" = None,
        encoding: str = "utf-8",
        dialect: "Optional[Union[csv.Dialect, str]]" = None,
        strict: bool = False,
        redefine: str = "raise",
        baseiri: "Optional[str]" = None,
        **kwargs,
    ) -> "TableDoc":
        # pylint: disable=line-too-long
        """Parse a csv file using the standard library csv module.

        Arguments:
            csvfile: Name of CSV file to parse or an iterable of strings.
            type: Type of data to save (applies to all rows).  Should
                either be one of the pre-defined names: "Dataset",
                "Distribution", "AccessService", "Parser" and "Generator"
                or an IRI to a class in an ontology.
            keywords: Keywords object with additional keywords definitions.
                If not provided, only default keywords are considered.
            context: Additional user-defined context that should be
                returned on top of the default context.  It may be a
                string with an URL to the user-defined context, a dict
                with the user-defined context or a sequence of strings
                and dicts.
            prefixes: Dict with prefixes in addition to those included in the
                JSON-LD context.  Should map namespace prefixes to IRIs.
            encoding: The encoding of the csv file.  Note that Excel may
                encode as "ISO-8859" (which was commonly used in the 1990th).
            dialect: A subclass of csv.Dialect, or the name of the dialect,
                specifying how the `csvfile` is formatted.  For more details,
                see [Dialects and Formatting Parameters].
            strict: Whether to raise an `InvalidKeywordError` exception if `d`
                contains an unknown key.
            redefine: Determine how to handle redefinition of existing
                keywords.  Should be one of the following strings:
                  - "allow": Allow redefining a keyword. Emits a
                    `RedefineKeywordWarning`.
                  - "skip": Don't redefine existing keyword. Emits a
                    `RedefineKeywordWarning`.
                  - "raise": Raise an RedefineError (default).
            baseiri: If given, it will be used as a base iri to
                resolve relative IRIs. (I.e. Not valid URLs).
            kwargs: Additional keyword arguments overriding individual
                formatting parameters.  For more details, see
                [Dialects and Formatting Parameters].

        Returns:
            New TableDoc instance.

        Note:
            For situations in which multiple resources are involved,
            see [multi-table workflows]:
            https://emmc-asbl.github.io/tripper/latest/datadoc/customisation/#multi-table-workflows

        References:
        [Dialects and Formatting Parameters]: https://docs.python.org/3/library/csv.html#dialects-and-formatting-parameters
        """

        def read(f, dialect):
            """Return csv reader from file-like object `f`."""
            if dialect is None and not kwargs:
                sample = f.read(1024)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
                except csv.Error:
                    # The build-in sniffer not always work well with
                    # non-numerical csv files. Try our simple sniffer
                    dialect = csvsniff(sample)
                finally:
                    f.seek(0)
            reader = csv.reader(f, dialect=dialect, **kwargs)
            headers = next(reader)
            data = list(reader)
            return headers, data

        if isinstance(csvfile, (str, Path)):
            with openfile(csvfile, mode="rt", encoding=encoding) as f:
                headers, data = read(f, dialect)
        else:
            headers, data = read(csvfile, dialect)

        return TableDoc(
            headers=headers,
            data=data,
            type=type,
            keywords=keywords,
            context=context,
            prefixes=prefixes,
            strict=strict,
            redefine=redefine,
            baseiri=baseiri,
        )

    def write_csv(
        self,
        csvfile: "Union[Path, str, Writer]",
        encoding: str = "utf-8",
        dialect: "Union[csv.Dialect, str]" = "excel",
        prefixes: "Optional[dict]" = None,
        **kwargs,
    ) -> None:
        # pylint: disable=line-too-long
        """Write the table to a csv file using the standard library csv module.

        Arguments:
            csvfile: File-like object or name of CSV file to write.
            encoding: The encoding of the csv file.
            dialect: A subclass of csv.Dialect, or the name of the dialect,
                specifying how the `csvfile` is formatted.  For more details,
                see [Dialects and Formatting Parameters].
            prefixes: Prefixes used to compact the headers.
            kwargs: Additional keyword arguments overriding individual
                formatting parameters.  For more details, see
                [Dialects and Formatting Parameters].

        References:
        [Dialects and Formatting Parameters]: https://docs.python.org/3/library/csv.html#dialects-and-formatting-parameters
        """

        def write(f):
            writer = csv.writer(f, dialect=dialect, **kwargs)

            # TODO: use self.context and compact to shortnames
            if prefixes:
                headers = []
                for h in self.headers:
                    for prefix, ns in prefixes.items():
                        if h.startswith(str(ns)):
                            headers.append(f"{prefix}:{h[len(str(ns)):]}")
                            break
                    else:
                        headers.append(h)
                writer.writerow(headers)
            else:
                writer.writerow(self.headers)

            for row in self.data:
                writer.writerow(row)

        if isinstance(csvfile, (str, Path)):
            with open(csvfile, mode="wt", encoding=encoding) as f:
                write(f)
        else:
            write(csvfile)

    @staticmethod
    def parse_excel(
        excelfile: "Union[Path, str]",
        sheet: "Union[str, int]" = 0,
        cellrange: "Optional[str]" = None,
        type: "Optional[str]" = None,
        keywords: "Optional[KeywordsType]" = None,
        context: "Optional[ContextType]" = None,
        prefixes: "Optional[dict]" = None,
        strip: bool = True,
        strict: bool = False,
        redefine: str = "raise",
        baseiri: "Optional[str]" = None,
    ) -> "TableDoc":
        # pylint: disable=line-too-long
        """Parse a csv file using the standard library csv module.

        Arguments:
            excelfile: Name of Excel file to parse.
            sheet: Sheet name or number to load.
            cellrange: Cell range to load. Examples: "A1:C4", "A:C", "1:4".
                The default is to read all cells.
            type: Type of data to save (applies to all rows).  Should
                either be one of the pre-defined names: "Dataset",
                "Distribution", "AccessService", "Parser" and "Generator"
                or an IRI to a class in an ontology.
            keywords: Keywords object with additional keywords definitions.
                If not provided, only default keywords are considered.
            context: Dict with user-defined JSON-LD context.
            prefixes: Dict with prefixes in addition to those included in the
                JSON-LD context.  Should map namespace prefixes to IRIs.
            strip: Whether to strip leading and trailing whitespaces from cells.
            strict: Whether to raise an `InvalidKeywordError` exception if `d`
                contains an unknown key.
            redefine: Determine how to handle redefinition of existing
                keywords.  Should be one of the following strings:
                  - "allow": Allow redefining a keyword. Emits a
                    `RedefineKeywordWarning`.
                  - "skip": Don't redefine existing keyword. Emits a
                    `RedefineKeywordWarning`.
                  - "raise": Raise an RedefineError (default).
            baseiri: If given, it will be used as a base iri to
                resolve relative IRIs. (I.e. Not valid URLs).

        Returns:
            New TableDoc instance.

        """
        # pylint: disable=import-outside-toplevel
        from openpyxl import load_workbook

        wb = load_workbook(
            excelfile,
            read_only=True,
            keep_vba=False,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )
        # Get worksheet
        ws = wb[wb.sheetnames[sheet] if isinstance(sheet, int) else sheet]

        # Get cell range
        if cellrange:
            cr = ws[cellrange]
        else:
            # Find first non-empty rows and columns
            nrows = next((i for i, r in enumerate(ws.values) if not r[0]), 0)
            ncols = next(
                (i for i, v in enumerate(next(ws.values)) if not v), 0
            )
            cr = ws.iter_rows(max_row=nrows, max_col=ncols)

        table = [[cell.value for cell in row] for row in cr]

        return TableDoc(
            headers=table[0],
            data=table[1:],
            type=type,
            keywords=keywords,
            context=context,
            prefixes=prefixes,
            strip=strip,
            strict=strict,
            redefine=redefine,
            baseiri=baseiri,
        )

    def unique_headers(self):
        """Return the headers with brackets appended to duplicated labels
        to make them unique.

        For example, the headers

            ["@id", "@type", "@type", "part.name", "part.name"]

        "distribution.downloadURL",

        would be renamed to

            ["@id", "@type[1]", "@type[2]", "part[1].name", "part[2].name"]

        """
        new = []
        seen = {}
        for h in self.headers:
            if self.headers.count(h) == 1:
                new.append(h)
            else:
                head, tail = h.split(".", 1) if "." in h else (h, None)
                n = seen[head] + 1 if head in seen else 1
                seen[head] = n
                new.append(f"{head}[{n}].{tail}" if tail else f"{head}[{n}]")
        return new


def csvsniff(sample):
    """Custom csv sniffer.

    Analyse csv sample and returns a csv.Dialect instance.
    """
    # Determine line terminator
    if "\r\n" in sample:
        linesep = "\r\n"
    else:
        counts = {s: sample.count(s) for s in "\n\r"}
        linesep = max(counts, key=lambda k: counts[k])

    lines = sample.split(linesep)
    del lines[-1]  # skip last line since it might be truncated
    if not lines:
        raise csv.Error(
            "too long csv header. No line terminator within sample"
        )
    headers = lines[0]

    # Possible delimiters and quote chars to check
    delims = [d for d in ",;\t :" if headers.count(d)]
    quotes = [q for q in "\"'" if sample.count(q)]
    if not quotes:
        quotes = ['"']

    # For each (quote, delim)-pair, count the number of tokens per line
    # Only pairs for which all lines has the same number of tokens are added
    # to ntokens
    ntokens = {}  # map (quote, delim) to number of tokens per line
    for q in quotes:
        for d in delims:
            ntok = []
            for ln in lines:
                # Remove quoted tokens
                ln = re.sub(f"(^{q}[^{q}]*{q}{d})|({d}{q}[^{q}]*{q}$)", d, ln)
                ln = re.sub(f"{d}{q}[^{q}]*{q}{d}", d * 2, ln)
                ntok.append(len(ln.split(d)))

            if ntok and max(ntok) == min(ntok):
                ntokens[(q, d)] = ntok[0]

    # From ntokens, select (quote, delim) pair that results in the highest
    # number of tokens per line
    if not ntokens:
        raise csv.Error("not able to determine delimiter")
    quote, delim = max(ntokens, key=lambda k: ntokens[k])

    class dialect(csv.Dialect):
        """Custom dialect."""

        # pylint: disable=too-few-public-methods
        _name = "sniffed"
        delimiter = delim
        doublequote = True  # quote chars inside quotes are duplicated
        # escapechar = "\\"  # unused
        lineterminator = linesep
        quotechar = quote
        quoting = csv.QUOTE_MINIMAL
        skipinitialspace = False  # don't ignore spaces before a delimiter
        strict = False  # be permissive on malformed csv input

    return dialect


class Column:
    """Help class representing a column."""

    # pylint: disable=too-few-public-methods

    def __init__(self, header, context=None, strip=True):
        # pylint: disable=line-too-long
        """Initialise a column opject.

        Arguments:
            header: Column header. For the allowed syntax of the header label
                `header`, see https://emmc-asbl.github.io/tripper/latest/datadoc/documenting-a-resource/#documenting-as-table
            context: Optional context. Used for deriving datatype.
            strip: Whether to strip white spaces from `header`.

        """
        if strip:
            header = header.strip()

        fields = re.findall(r"([^.\[]+)(\[([^\]]*)\])?", header)
        label = fields[0][2].split("?", 1)[0]
        spec = fields[-1][2].split("?", 1)

        options = {}
        if len(spec) == 2:
            for opt in spec[1].split("&"):
                k, v = opt.split("=", 1)
                options[k] = v

        datatype = None
        leafname = fields[-1][0]
        if context and not leafname.startswith("@"):
            df = context.getdef(leafname)
            if "@type" in df and df["@type"] != "@id":
                datatype = df["@type"]

        self.header = header
        self.context = context
        self.strip = strip
        self.names = [f[0] for f in fields]
        self.label = label
        self.options = options
        self.datatype = datatype

    def add(self, d, cell):
        """Add cell value to dict `d`."""
        if not cell:
            return
        if "sep" in self.options:
            vals = cell.split(self.options["sep"])
        else:
            vals = [cell]

        for v in vals:
            val = (
                Literal(v, datatype=self.datatype).value
                if self.datatype
                else v
            )
            # if "unit" in self.options:
            #    val = {
            #        "value": val,
            #        "unit": self.unit,
            #    }
            addnested(d, self.header, val)
